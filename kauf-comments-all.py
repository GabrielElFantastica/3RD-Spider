# -*- coding = utf-8 -*-
# @Time : 2022-04-02 15:48
# @Author : 帅哥张
# @File : kauf-comments-all.py
# @Software: PyCharm
import json

import execjs
import requests
import openpyxl
import pandas as pd

class Py4Js:

    def __init__(self):
        self.ctx = execjs.compile(""" 
        function TL(a) { 
        var k = ""; 
        var b = 406644; 
        var b1 = 3293161072; 

        var jd = "."; 
        var $b = "+-a^+6"; 
        var Zb = "+-3^+b+-f"; 

        for (var e = [], f = 0, g = 0; g < a.length; g++) { var m = a.charCodeAt(g); 128 > m ? e[f++] = m : (2048 > m 
        ? e[f++] = m >> 6 | 192 : (55296 == (m & 64512) && g + 1 < a.length && 56320 == (a.charCodeAt(g + 1) & 64512) 
        ? (m = 65536 + ((m & 1023) << 10) + (a.charCodeAt(++g) & 1023), e[f++] = m >> 18 | 240, e[f++] = m >> 12 & 63 
        | 128) : e[f++] = m >> 12 | 224, e[f++] = m >> 6 & 63 | 128), e[f++] = m & 63 | 128) } a = b; for (f = 0; f < 
        e.length; f++) a += e[f], a = RL(a, $b); a = RL(a, Zb); a ^= b1 || 0; 0 > a && (a = (a & 2147483647) + 
        2147483648); a %= 1E6; return a.toString() + jd + (a ^ b) }; 

    function RL(a, b) { 
        var t = "a"; 
        var Yb = "+"; 
        for (var c = 0; c < b.length - 2; c += 3) { 
            var d = b.charAt(c + 2), 
            d = d >= t ? d.charCodeAt(0) - 87 : Number(d), 
            d = b.charAt(c + 1) == Yb ? a >>> d: a << d; 
            a = b.charAt(c) == Yb ? a + d & 4294967295 : a ^ d 
        } 
        return a 
    } 
    """)

    def get_tk(self, text):
        return self.ctx.call("TL", text)


# 谷歌翻译方法
def google_translate(content):
    """实现谷歌的翻译"""
    js = Py4Js()
    tk = js.get_tk(content)

    if len(content) > 4891:
        print("翻译的长度超过限制！！！")
        return

    param = {'tk': tk, 'q': content}

    result = requests.get("http://translate.google.com/translate_a/single?client=t&sl=de&tl=zh-CN&dt=t", params=param)

    # 返回的结果为Json，解析为一个嵌套列表
    trans = result.json()[0]
    # print(trans)
    ret = ''
    for i in range(len(trans)):
        line = trans[i][0]
        if line is not None:
            ret += trans[i][0]

    return ret

# file = pd.read_excel(io="Kaufland.xls")
# print(file)
# data = file.loc[:, 'text']
# for da in list(data):
#     print(type(da))
#     try:
#         da = google_translate(da)
#     except json.decoder.JSONDecodeError:
#         print("空值")
#         continue
#     print(da)
# print(data)

def get_translation(text):
    try:
        text = google_translate(text)
    except json.decoder.JSONDecodeError:
        text = ""
    return text


wb = openpyxl.load_workbook("Attribut翻译汇总.xlsx")
# ws = wb['太阳伞']
categories = wb.sheetnames
# categories = ["休闲椅"]
print(categories)
for category in categories:
    ws = wb[category]
    data = []
    for i in range(3, len(ws["A"]) + 1):
        if not ws['A%d' % i].value:
            continue
        elif ws['B%d' % i].value:
            continue
        else:
            pass

        try:
            translated = google_translate(ws['A%d' % i].value)
        except Exception:
            translated = ""
            continue
        finally:
            ws['B%d' % i].value = translated
            data.append(translated)
            print(translated)


    print(data)

wb.save("Attribut翻译汇总.xlsx")



