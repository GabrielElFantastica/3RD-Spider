# -*- coding = utf-8 -*-
# @Time : 2022-03-03 11:23
# @Author : 帅哥张
# @File : kauf-rating.py
# @Software: PyCharm
import datetime
import pandas as pd
import requests
import openpyxl
from openpyxl.styles import Font
from bs4 import BeautifulSoup  # 解析网页返回的html文件的包
import re  # 使用正则表达式的包
import time

# 正则表达式，用于定义所需字符串在页面html文件中的位置，具体可使用网页版的开发者工具进行查看
# findLink = re.compile(r'<a.*href="(.*?)".*>')
findTitle = re.compile(r'<div class="product__title">(.*?)</div>', re.S)  # 匹配标题
findQa = re.compile(r'class="product__sponsored-ad-label">.*(.*?).*</div>', re.S)  # 页面排名
findID = re.compile(r'/product/(\d*)/', re.S)  # kaufland专用id
findSeller = re.compile(r'(.*?) ', re.S)  # 竞卖
findSKU = re.compile(r'.* (.*?)$', re.S)  # SKU
findCategory = re.compile(r'<a.*data-category-name="(.*?)"')  # kaufland内置类目
findComment = re.compile(r'class="product-rating__count">\n        (\d*)\n', re.S)  # 评价数


def getData(baseurl):
    i, b = 0, 0
    for a in range(1, 10):  # 开始循环，从第几页到第几页
        if leixing == "ProductList":
            url = baseurl + "p" + str(a) + "/"  # 网址
        else:
            url = baseurl + "&page=" + str(a)
        print(url)
        head = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/93.0.4577.63 Safari/537.36 "
        }
        html = requests.get(url=url, headers=head).text  # 启动定义好的askurl程序，url在上边设定好了
        bs = BeautifulSoup(html, features="lxml")  # 解析网页，将html文件变成txt
        items = bs.find_all('article')  # 找到所有<article>，锁定
        for item in items:  # 在所有article中循环每一个article
            i += 1  # 每次循环+1
            print("第%s条" % i)
            # print(item)
            data = [str(datetime.date.today())]  # 清空data，重新循环
            item = str(item)  # 把item从集合中的元素转化为字符串
            # print(item)
            # 逐个解析

            title = re.findall(findTitle, item)
            # print(title)
            title = title[0]
            title = title.strip()  # 去掉前后的空格
            # print(title)

            seller = re.findall(findSeller, title)[0]
            sku = re.findall(findSKU, title)[0]
            # print(seller)
            # if seller not in ["VASAGLE", "SONGMICS", "FEANDREA"]:
            #     continue

            data.append(suche)
            data.append(leixing)

            if seller in ['VASAGLE', 'SONGMICS', 'FEANDREA']:
                data.append(seller)
                data.append(sku)
            elif seller in ['Vicco', 'VICCO', 'SoBuy®', 'WOLTU', 'vidaXL', 'HOMEXPERTS']:
                data.append(seller)
                data.append("")
            else:
                data.append("")
                data.append("")

            data.append(title)

            try:
                productid = re.findall(findID, item)
                # print(id)
                data.append(productid[0])
            except Exception:
                data.append("")

            qa = re.findall(findQa, item)  # 根据上边写的正则表达式，在item中进行匹配，获得的是一个列表
            # print(qa)
            try:
                # qa = int(qa[0])
                qa = qa[0]
                data.append(qa)  # 取所有找到的qa中的第一个
                data.append("sor")
            except IndexError:
                b += 1
                data.append(b)
                data.append("")

            comment = re.findall(findComment, item)
            try:
                data.append(int(comment[0]))
            except IndexError:
                data.append("")

            print(data)  # 看看每一个item中得到的data有没有错误
            # 添加数据
            datalist.append(data)  # 把data添加到datalist中，然后开始下次循环

    # print(datalist)
    return datalist  # 获得datalist


def saveData(savepath):
    print("save....")
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'zhang'
    style = Font(name='Calibri', size=12)
    col = ("日期", "类目", "类型", "卖家", "SKU", "标题", "ID", "排名", "广告标志", "评价数")
    for i in range(0, len(col)):
        # print(col[i])
        sheet.cell(row=1, column=i + 1).value = col[i]
    print("写好标题啦")
    for a in range(0, len(datalist)):
        print("第%d条" % (a + 1))
        for b in range(0, len(datalist[0])):
            # print(datalist[a][b])
            sheet.cell(row=a + 2, column=b + 1).value = datalist[a][b]
            sheet.cell(row=a + 2, column=b + 1).font = style
    wb.save(savepath)  # 保存
    print("文件保存完毕")


if __name__ == "__main__":
    start = time.localtime()
    datalist = []
    file = pd.read_excel("kauf-rating.xlsx", sheet_name="关键词")  # 放网址的表格
    print(file)
    categories = list(file.loc[:, 'Category'])
    print(len(categories))
    links = list(file.loc[:, 'Link'])
    print(links)
    leixings = list(file.loc[:, '类型'])
    print(leixings)
    savepath = "KAUF-Ranking-%s.xlsx" % datetime.date.today()

    for a in range(0, len(categories)):
        suche = categories[a]
        leixing = leixings[a]
        baseurl = links[a]
        try:
            getData(baseurl)
        except Exception as e:
            print("半路出错")
            if hasattr(e, "code"):  # 错误代码
                print(e.code)
            if hasattr(e, "reason"):  # 错误原因
                print(e.reason)
            pass
    saveData(savepath)
    end = time.localtime()
    print("开始时间：%s" % (time.strftime("%Y-%m-%d %H:%M:%S", start)))
    print("结束时间：%s" % (time.strftime("%Y-%m-%d %H:%M:%S", end)))
